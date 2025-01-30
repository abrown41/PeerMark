from teams import get_teams, extract_teams
from openpyxl import load_workbook

import sys
import os
# csv file containing list of students and their teams
teams_file = "./teams.csv"
# xlsx file containing template spreadsheet
template_spreadsheet = "./template_spreadsheet.xlsx"

# Check input files all exist
for fname in [teams_file, template_spreadsheet]:
    if (not os.path.isfile(fname)):
        sys.exit("File " + fname + " not found")


def write_sheet(fname, team, wb):
    # Construct student spreadsheet containing the necessary student names
    from openpyxl import load_workbook
    from openpyxl.styles import Protection

    opname = fname+"_peer_review.xlsx"
    wb.save(filename=opname)

    opwb = load_workbook(opname)
    opws = opwb.active
    opws.title = "Peer Review"
    opws.protection.sheet = True

    for col, name in enumerate(team, 2):
        opws.cell(1, col, name)
        for row in [2, 3, 4, 5, 6, 7, 9, 11]:
            opws.cell(row, col).protection = Protection(locked=False)

    opwb.save(opname)


wb = load_workbook(template_spreadsheet)
teamlist = extract_teams(teams_file)
stud_names = (get_teams(teams_file))

for team in teamlist:
    try:
        stud_list = stud_names[team]
    except KeyError:
        stud_list = False
        pass

    if stud_list:
        teamnames = [name[0] for name in stud_list]
#        for name in stud_list:
#            und_name = name[0].replace(" ", "_")+"_"+name[1]
        tname = team.replace(" ", "_")
        write_sheet(tname, teamnames, wb)
